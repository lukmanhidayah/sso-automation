import json
import os
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
from typing import Any, Dict, List
import ijson
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from selected_no_peserta import selected_no_peserta

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
except Exception:
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore


API_URL = (
    "https://api-siasn.bkn.go.id/siasn-instansi/pengadaan/usulan/monitoring"
    "?no_peserta=&nama=&tgl_usulan=&jenis_pengadaan_id=02&jenis_formasi_id=&status_usulan=&periode=2024&limit=100000&offset=0"
)

# Mapping of status_usulan ID to descriptive name
STATUS_USULAN_MAP = {
    "1": "Input Berkas",
    "2": "Berkas Disimpan (Terverifikasi)",
    "3": "Surat Usulan",
    "4": "Approval Surat Usulan",
    "5": "Perbaikan Dokumen",
    "6": "Tidak Memenuhi Syarat",
    "7": "Menunggu Cetak SK – Menyetujui",
    "8": "Menunggu Cetak SK – Perbaikan Pertek",
    "9": "Menunggu Cetak SK – Pembatalan Pertek",
    "10": "Cetak SK",
    "11": "Profil PNS telah diperbaharui",
    "12": "Terima Usulan",
    "13": "Validasi Usulan – Tidak Memenuhi Syarat",
    "14": "Validasi Usulan – Perbaikan Dokumen",
    "15": "Validasi Usulan – Disetujui",
    "16": "Berkas Disetujui",
    "17": "Menunggu Paraf – Paraf Pertek",
    "18": "Menunggu Paraf – Gagal Paraf Pertek",
    "19": "Sdh di paraf - Pertek",
    "20": "Menunggu Tanda tangan- TTD Pertek",
    "21": "Berkas Ditolak - TTD Pertek",
    "22": "Sdh di TTD - Pertek",
    "23": "Surat Keluar",
    "24": "Perbaikan Pertek (Menunggu Approval Instansi)",
    "25": "Terima Usulan Penetapan – Pembatalan",
    "26": "Pembatalan Pertek (Menunggu Approval Instansi)",
    "27": "Menunggu SK – Paraf / TTE",
    "28": "Setuju Paraf SK",
    "29": "Tolak TTD SK",
    "30": "Setuju TTD SK",
    "31": "Telah Update di Profile PNS",
    "32": "Pembuatan SK Berhasil",
    "33": "Menunggu Layanan",
    "34": "Perbaikan Dokumen - Menunggu Approval",
    "35": "Tolak Paraf SK",
    "36": "Menunggu TTD - SK",
    "37": "Approval Perbaikan Pertek",
    "38": "Approval Pembatalan Pertek",
    "39": "Perbaikan SK",
    "40": "Berkas Disimpan (Terverifikasi) - Perbaikan SK",
    "41": "Validasi Usulan - Perbaikan SK",
    "42": "Validasi Usulan - Perbaikan SK (Disetujui)",
    "43": "Menunggu Paraf - Perbaikan SK",
    "44": "Menunggu TTD - Perbaikan SK",
    "45": "Sudah TTD - Perbaikan SK",
    "46": "Menunggu TTD SK - Instansi",
    "47": "Tolak TTD SK - Instansi",
    "48": "Setuju TTD SK - Instansi",
    "49": "Sudah TTD - SK",
    "50": "Perbaikan Dokumen - MYSAPK",
    "51": "Input Berkas - Perbaikan MySAPK ",
    "52": "Perbaikan Dokumen - Approval",
    "53": "Setuju TTD Pertek",
    "55": "Approval Tingkat Provinsi",
    "56": "Perbaikan Approval",
    "57": "Perbaikan Pertek",
    "58": "Validasi Usulan - Perbaikan Pertek",
    "59": "Menunggu Buat Sk",
    "60": "Proses Persidangan",
    "61": "Input Berkas - SK PNS",
    "62": "Menunggu TTD SK PNS - Instansi",
    "63": "Setuju TTD Digital SK PNS",
    "64": "Pembuatan SK Basah PNS Berhasil",
    "65": "Pembatalan NIP/Pertek",
    "66": "Perbaikan SK Provinsi",
    "67": "Validasi Perbaikan Dokumen - BTS",
    "68": "Validasi Perbaikan SK - BTS",
    "69": "Perbaikan Dokumen SK - BTS",
    "70": "Tolak Perbaikan SK - TMS",
    "71": "Validasi Perbaikan Dokumen SK - BTS",
    "72": "Perbaikan Dokumen Pertek - BTS",
    "73": "Tolak Perbaikan Pertek - TMS",
    "74": "Validasi Perbaikan Pertek - BTS",
    "75": "Pembatalan PERTEK",
    "76": "Menunggu Rekom OTDA",
    "99": "Usulan Dihapus",
}


def load_sso_token(path: str) -> str:
    if not os.path.exists(path):
        raise FileNotFoundError(f"LocalStorage JSON not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    token = data.get("sso_token")
    if not token:
        raise ValueError("Key 'sso_token' not found in localStorage JSON")
    return token


def download_monitoring_usulan(
    out_path: str, localstorage_path: str = "data/sso_localstorage.json"
) -> None:
    print("Downloading monitoring_usulan data...")
    token = load_sso_token(localstorage_path)

    headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9,id;q=0.8",
        "Authorization": f"Bearer {token}",
        "Connection": "keep-alive",
        "Origin": "https://siasn-instansi.bkn.go.id",
        "Referer": "https://siasn-instansi.bkn.go.id/layananPengadaan/monitoringUsulan",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/139.0.0.0 Safari/537.36 Edg/139.0.0.0"
        ),
        "sec-ch-ua": '"Not;A=Brand";v="99", "Microsoft Edge";v="139", "Chromium";v="139"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        # Cookie header is optional; include if needed
        # "Cookie": "d732bf4cc14bde954e7b20b82e794606=730362334080c57616252771951a529f",
    }

    req = Request(API_URL, headers=headers, method="GET")
    try:
        with urlopen(req) as resp:
            status = resp.getcode()
            body = resp.read()
            if status != 200:
                raise HTTPError(API_URL, status, f"HTTP {status}", resp.headers, None)
    except HTTPError as e:
        # Try to read error body for debugging
        detail = None
        try:
            detail = e.read().decode("utf-8", errors="ignore")  # type: ignore[attr-defined]
        except Exception:
            pass
        msg = f"Request failed: {e.code} {e.reason}"
        if detail:
            msg += f"\n{detail}"
        raise RuntimeError(msg)
    except URLError as e:
        raise RuntimeError(f"Network error: {e.reason}")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "wb") as f:
        f.write(body)


def download_monitoring_usulan_paginated(
    out_path: str,
    localstorage_path: str = "data/sso_localstorage.json",
    per_page: int = 10000,
) -> None:
    print("Downloading monitoring_usulan data with pagination...")
    token = load_sso_token(localstorage_path)

    headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9,id;q=0.8",
        "Authorization": f"Bearer {token}",
        "Connection": "keep-alive",
        "Origin": "https://siasn-instansi.bkn.go.id",
        "Referer": "https://siasn-instansi.bkn.go.id/layananPengadaan/monitoringUsulan",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/139.0.0.0 Safari/537.36 Edg/139.0.0.0"
        ),
        "sec-ch-ua": '"Not;A=Brand";v="99", "Microsoft Edge";v="139", "Chromium";v="139"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write('{"data":[')
        offset = 0
        total = None
        first = True
        while True:
            url = (
                "https://api-siasn.bkn.go.id/siasn-instansi/pengadaan/usulan/monitoring"
                f"?no_peserta=&nama=&tgl_usulan=&jenis_pengadaan_id=02&jenis_formasi_id=&status_usulan=&periode=2024"
                f"&limit={per_page}&offset={offset}"
            )
            req = Request(url, headers=headers, method="GET")
            try:
                with urlopen(req) as resp:
                    status = resp.getcode()
                    body = resp.read()
                    if status != 200:
                        raise HTTPError(
                            url, status, f"HTTP {status}", resp.headers, None
                        )
            except HTTPError as e:
                detail = None
                try:
                    detail = e.read().decode("utf-8", errors="ignore")  # type: ignore[attr-defined]
                except Exception:
                    pass
                msg = f"Request failed: {e.code} {e.reason}"
                if detail:
                    msg += f"\n{detail}"
                raise RuntimeError(msg)
            except URLError as e:
                raise RuntimeError(f"Network error: {e.reason}")

            resp_json = json.loads(body)
            if total is None:
                total = resp_json.get("meta", {}).get("total", 0)
                print(f"Total data: {total}")
            page_data = resp_json.get("data", [])
            print(f"Fetched {len(page_data)} items (offset {offset})")

            for item in page_data:
                if not first:
                    f.write(",")
                f.write(json.dumps(item, ensure_ascii=False))
                first = False

            if len(page_data) < per_page:
                break
            offset += per_page
            time.sleep(1)
        f.write("]}")
    print(f"Saved all data to {out_path}")


def convert_monitoring_json_to_excel(
    json_path: str,
    excel_path: str,
    pertek_drive_folder_id: str | None = None,
) -> None:
    """Convert monitoring_usulan JSON into an Excel file dengan No. Peserta terpilih.

    Jika `pertek_drive_folder_id` diberikan, cek file Pertek yang sudah
    ada di Google Drive (judul file) dan isi kolom "Drive URL" otomatis.
    """
    print("Converting JSON to Excel (streaming)...")
    if Workbook is None:
        raise RuntimeError(
            "openpyxl not available. Please install it (e.g., pip install openpyxl)."
        )

    if not os.path.exists(json_path):
        raise FileNotFoundError(f"JSON file not found: {json_path}")

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    # Siapkan peta judul->link dari folder Drive berisi PDF Pertek (opsional)
    drive_title_link_map: dict[str, str] = {}
    if pertek_drive_folder_id:
        try:
            from drive_upload import list_title_to_link_map  # type: ignore
            drive_title_link_map = list_title_to_link_map(pertek_drive_folder_id)
        except Exception as e:
            print(f"Peringatan: gagal memuat daftar file Pertek dari Drive: {e}")
            drive_title_link_map = {}
    wb = Workbook()
    ws = wb.active
    ws.title = "monitoring_usulan"
    ws.append(["No. Peserta", "NIP", "Nama", "Status Usulan", "Drive URL"])
    for col_idx in range(1, 6):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 50

    processed_no_peserta = set()  # Set untuk melacak unik
    missing_count = 0

    with open(json_path, "r", encoding="utf-8") as f:
        for it in ijson.items(f, "data.item"):
            nested = (it or {}).get("usulan_data") or {}
            nested_data = nested.get("data") or {}
            no_peserta = nested_data.get("no_peserta") or ""
            if no_peserta not in selected_no_peserta:
                continue
            if no_peserta in processed_no_peserta:
                print(f"Duplikasi dilewati: {no_peserta}")
                continue  # Lewati duplikasi
            processed_no_peserta.add(no_peserta)
            nama = (it or {}).get("nama") or nested_data.get("nama") or ""
            status_usulan = (it or {}).get("status_usulan") or ""
            status_id = str(status_usulan)
            status_usulan_name = STATUS_USULAN_MAP.get(status_id, status_id)
            nip = (it or {}).get("nip") or ""
            title_base = _sanitize_filename(
                f"Pertek_{nip}_{nama}" if (nip and nama) else (f"Pertek_{nip}" if nip else "")
            )
            drive_url = drive_title_link_map.get(title_base, "") if title_base else ""
            ws.append([no_peserta, nip, nama, status_usulan_name, drive_url])
            if not no_peserta:
                missing_count += 1

    missing_no_peserta = selected_no_peserta - processed_no_peserta
    print(f"Total no_peserta tidak ditemukan: {len(missing_no_peserta)}")

    # Load token for additional requests
    localstorage_path = "data/sso_localstorage.json"  # Adjust if needed
    token = load_sso_token(localstorage_path)

    headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9,id;q=0.8",
        "Authorization": f"Bearer {token}",
        "Connection": "keep-alive",
        "Origin": "https://siasn-instansi.bkn.go.id",
        "Referer": "https://siasn-instansi.bkn.go.id/layananPengadaan/monitoringUsulan",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/139.0.0.0 Safari/537.36 Edg/139.0.0.0"
        ),
        "sec-ch-ua": '"Not;A=Brand";v="99", "Microsoft Edge";v="139", "Chromium";v="139"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    # Tampung item baru yang ditemukan saat pencarian untuk ditambahkan ke JSON
    new_items_to_append: list[dict] = []

    for no_peserta in missing_no_peserta:
        print(f"Mencari data untuk no_peserta: {no_peserta}")
        url = (
            "https://api-siasn.bkn.go.id/siasn-instansi/pengadaan/usulan/monitoring"
            f"?no_peserta={no_peserta}&nama=&tgl_usulan=&jenis_pengadaan_id=02&jenis_formasi_id=&status_usulan=&periode=2024&limit=1&offset=0"
        )
        req = Request(url, headers=headers, method="GET")
        try:
            with urlopen(req) as resp:
                status = resp.getcode()
                body = resp.read()
                if status != 200:
                    raise HTTPError(url, status, f"HTTP {status}", resp.headers, None)
        except HTTPError as e:
            detail = None
            try:
                detail = e.read().decode("utf-8", errors="ignore")  # type: ignore[attr-defined]
            except Exception:
                pass
            msg = f"Request failed for {no_peserta}: {e.code} {e.reason}"
            if detail:
                msg += f"\n{detail}"
            print(msg)
            continue
        except URLError as e:
            print(f"Network error for {no_peserta}: {e.reason}")
            continue

        resp_json = json.loads(body)
        page_data = resp_json.get("data", [])
        if page_data:
            item = page_data[0]
            nested = (item or {}).get("usulan_data") or {}
            nested_data = nested.get("data") or {}
            nama = (item or {}).get("nama") or nested_data.get("nama") or ""
            status_usulan = (item or {}).get("status_usulan") or ""
            status_id = str(status_usulan)
            status_usulan_name = STATUS_USULAN_MAP.get(status_id, status_id)
            nip = (item or {}).get("nip") or ""
            title_base = _sanitize_filename(
                f"Pertek_{nip}_{nama}" if (nip and nama) else (f"Pertek_{nip}" if nip else "")
            )
            drive_url = drive_title_link_map.get(title_base, "") if title_base else ""
            ws.append([no_peserta, nip, nama, status_usulan_name, drive_url])
            print(f"Data ditemukan dan ditambahkan untuk {no_peserta}")
            # Simpan item untuk ditambahkan ke monitoring_usulan.json
            if isinstance(item, dict):
                new_items_to_append.append(item)
        else:
            # Tanpa data item, tidak punya NIP/Nama untuk menebak judul Pertek -> kosongkan
            ws.append([no_peserta, "", "Tidak Ditemukan", "Tidak Ditemukan", ""])  # no link
            print(f"Data masih tidak ditemukan untuk {no_peserta}")
        time.sleep(1)  # Delay to avoid rate limiting

    # Setelah semua pencarian selesai, tambahkan item yang ditemukan ke JSON sumber
    # Lakukan dengan merge streaming agar tidak memakan memori besar (menghindari container mati)
    if new_items_to_append:
        try:
            temp_path = json_path + ".tmp"
            existing_ids: set[str] = set()
            existing_np: set[str] = set()

            # Tulis ulang file JSON secara streaming + sisipkan item baru di akhir
            with open(temp_path, "w", encoding="utf-8") as out_f:
                out_f.write('{"data":[')
                first = True

                # Baca data lama secara streaming dan salin apa adanya (dump kembali)
                try:
                    with open(json_path, "r", encoding="utf-8") as in_f:
                        for it in ijson.items(in_f, "data.item"):
                            if not isinstance(it, dict):
                                continue
                            # Kumpulkan kunci unik untuk deduplikasi
                            idv = (
                                (it.get("id") or "").strip()
                                if isinstance(it.get("id"), str)
                                else str(it.get("id")) if it.get("id") is not None else ""
                            )
                            nested = (it.get("usulan_data") or {}) if isinstance(it.get("usulan_data"), dict) else {}
                            nested_data = (nested.get("data") or {}) if isinstance(nested.get("data"), dict) else {}
                            npv = (nested_data.get("no_peserta") or "").strip()
                            if idv:
                                existing_ids.add(idv)
                            if npv:
                                existing_np.add(npv)

                            if not first:
                                out_f.write(",")
                            out_f.write(json.dumps(it, ensure_ascii=False))
                            first = False
                except (json.JSONDecodeError, FileNotFoundError):
                    # Jika file belum ada/korup, kita akan memulai dari nol di bawah
                    pass

                # Tambahkan item baru yang belum ada
                appended = 0
                for it in new_items_to_append:
                    if not isinstance(it, dict):
                        continue
                    idv = (
                        (it.get("id") or "").strip()
                        if isinstance(it.get("id"), str)
                        else str(it.get("id")) if it.get("id") is not None else ""
                    )
                    nested = (it.get("usulan_data") or {}) if isinstance(it.get("usulan_data"), dict) else {}
                    nested_data = (nested.get("data") or {}) if isinstance(nested.get("data"), dict) else {}
                    npv = (nested_data.get("no_peserta") or "").strip()
                    if (idv and idv in existing_ids) or (npv and npv in existing_np):
                        continue
                    if not first:
                        out_f.write(",")
                    out_f.write(json.dumps(it, ensure_ascii=False))
                    first = False
                    if idv:
                        existing_ids.add(idv)
                    if npv:
                        existing_np.add(npv)
                    appended += 1

                out_f.write("]}")

            # Ganti file asli dengan yang baru secara atomik
            os.replace(temp_path, json_path)
            if appended:
                print(f"Ditambahkan {appended} item baru ke {json_path}")
        except Exception as e:
            print(f"Gagal menambahkan item ke JSON (stream-merge): {e}")

    print(f"Total item diproses: {len(processed_no_peserta)}")
    print(f"Item dengan no_peserta kosong: {missing_count}")
    wb.save(excel_path)


def _sanitize_filename(name: str) -> str:
    """Sanitize filename by removing/replacing invalid characters for most OS."""
    invalid = '<>:"/\\|?*\n\r\t'
    sanitized = ''.join(c if c not in invalid else '_' for c in name)
    # Also strip leading/trailing spaces and dots commonly problematic on Windows
    return sanitized.strip(' .')


def download_pertek_documents_from_json(
    json_path: str,
    out_dir: str = "data/downloads/monitoring_usulan_ttd_pertek",
    localstorage_path: str = "data/sso_localstorage.json",
    excel_path: str | None = None,
    drive_folder_id: str | None = None,
    max_workers: int | None = None,
) -> None:
    """
    Read monitoring_usulan JSON and download each available Pertek by ID only
    as a PDF from the SIASN document endpoint, using the SSO token.

    Only downloads entries whose `no_peserta` is in `selected_no_peserta`.

    Output filenames follow: Pertek_{nip}_{nama}.pdf
    """
    print("Downloading Pertek documents from JSON...")
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"JSON file not found: {json_path}")

    token = load_sso_token(localstorage_path)

    base_pertek_url = (
        "https://api-siasn.bkn.go.id/siasn-instansi/pengadaan/dokumen/pertek/"
    )
    headers = {
        "Accept": "application/pdf,application/octet-stream;q=0.9,*/*;q=0.8",
        "Authorization": f"Bearer {token}",
        "Origin": "https://siasn-instansi.bkn.go.id",
        "Referer": "https://siasn-instansi.bkn.go.id/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/139.0.0.0 Safari/537.36 Edg/139.0.0.0"
        ),
    }

    os.makedirs(out_dir, exist_ok=True)

    # Concurrency settings
    if max_workers is None:
        try:
            max_workers = int(os.getenv("PERTEK_WORKERS", "10"))
        except ValueError:
            max_workers = 10

    # Build task list from JSON (only selected participants)
    tasks: List[Dict[str, str]] = []
    skipped = 0
    with open(json_path, "r", encoding="utf-8") as f:
        for it in ijson.items(f, "data.item"):
            if not isinstance(it, dict):
                continue
            item_id = (it.get("id") or "").strip()
            if not item_id:
                skipped += 1
                continue
            nip = (it.get("nip") or "").strip()
            nama = (it.get("nama") or "").strip()
            nested = (it or {}).get("usulan_data") or {}
            nested_data = nested.get("data") or {}
            no_peserta = (nested_data.get("no_peserta") or "").strip()

            if not no_peserta or (no_peserta not in selected_no_peserta):
                skipped += 1
                continue

            if not nip:
                nip = item_id
            fname_base = _sanitize_filename(
                f"Pertek_{nip}_{nama}" if nama else f"Pertek_{nip}"
            )
            out_file = os.path.join(out_dir, f"{fname_base}.pdf")
            tasks.append(
                {
                    "item_id": item_id,
                    "no_peserta": no_peserta,
                    "out_file": out_file,
                }
            )

    print(f"Total Pertek tasks: {len(tasks)} | skipped (filtered): {skipped}")

    # Worker function for parallel download + optional Drive upload
    def _worker(task: Dict[str, str]) -> Dict[str, str]:
        item_id = task["item_id"]
        no_peserta = task["no_peserta"]
        out_file = task["out_file"]

        url = base_pertek_url + item_id
        req = Request(url, headers=headers, method="GET")
        body = None
        last_error = None
        try:
            with urlopen(req) as resp:
                status = resp.getcode()
                data = resp.read()
                if status == 200 and data:
                    body = data
                else:
                    last_error = f"HTTP {status}"
        except HTTPError as e:
            try:
                _ = e.read().decode("utf-8", errors="ignore")  # type: ignore[attr-defined]
            except Exception:
                pass
            last_error = f"{e.code} {e.reason}"
        except URLError as e:
            last_error = f"Network error: {e.reason}"

        if body is None:
            print(f"Gagal download Pertek untuk {no_peserta} | {last_error}")
            return {"no_peserta": no_peserta, "saved": "0", "drive_url": ""}

        try:
            with open(out_file, "wb") as of:
                of.write(body)
            print(f"Saved: {out_file}")
            time.sleep(0.1)
        except Exception as e:
            print(f"Failed saving file {out_file}: {e}")
            return {"no_peserta": no_peserta, "saved": "0", "drive_url": ""}

        drive_url = ""
        if drive_folder_id:
            try:
                from drive_upload import upload_file_to_drive  # type: ignore

                title = os.path.splitext(os.path.basename(out_file))[0]
                drive_url = upload_file_to_drive(
                    out_file,
                    drive_folder_id,
                    convert_spreadsheet=False,
                    replace_by_title=True,
                    custom_title=title,
                )
            except Exception as e:
                print(f"Gagal upload ke Drive untuk {out_file}: {e}")
                drive_url = ""

        return {"no_peserta": no_peserta, "saved": "1", "drive_url": drive_url}

    # Run tasks in parallel
    downloaded = 0
    results: List[Dict[str, str]] = []
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = [ex.submit(_worker, t) for t in tasks]
        for fut in as_completed(futures):
            try:
                res = fut.result()
                results.append(res)
                if res.get("saved") == "1":
                    downloaded += 1
            except Exception as e:
                print(f"Task error: {e}")

    # Update Excel once at the end to avoid concurrent writes
    if excel_path and load_workbook is not None and os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
            col_np = header_map.get("No. Peserta", 1)
            col_url = header_map.get("Drive URL", ws.max_column + 1)
            if col_url > ws.max_column:
                ws.cell(row=1, column=col_url, value="Drive URL")
            url_map = {r["no_peserta"]: r.get("drive_url", "") for r in results if r.get("saved") == "1"}
            if url_map:
                for r in range(2, ws.max_row + 1):
                    np_val = str(ws.cell(row=r, column=col_np).value or "").strip()
                    if np_val in url_map:
                        ws.cell(row=r, column=col_url, value=url_map[np_val])
            wb.save(excel_path)
        except Exception as e:
            print(f"Gagal update Excel (batch): {e}")

    print(
        f"Pertek download complete. Downloaded: {downloaded}, Skipped: {skipped}, Failed: {len(tasks) - downloaded}"
    )
